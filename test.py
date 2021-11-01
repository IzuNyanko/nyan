import math
import os
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, LineChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.layout import Layout, ManualLayout # グラフの大きさ、位置調整用
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font, RichTextProperties
from openpyxl.drawing.line import LineProperties


from openpyxl.xml.constants import MAX_COLUMN
import pandas as pd
import numpy as np

# リファクタリングお願いします

# 定数
chartHeight = 16
chartWidth = 24
i = 1
xlFileName = input("pls input xlFileName\n")
wb = Workbook()
wb.save(xlFileName + ".xlsx")
tsvFolderName = input("pls input folder path\n")
fileList = os.listdir(tsvFolderName)
for tsvFileName in fileList:
    if "Out.txt" not in tsvFileName:
        continue
    #xlFileName = "test"
    #tsvFileName = "GaN_InN5nm1V_AlN_Out.txt"

    # Y, Ec, Ev, Ef, nの列のみを読み込む
    # なんかよく分からないけどpandas使ってます(python内でデータ処理するかもしれないので)
    data = pd.read_table(tsvFolderName + "/" + tsvFileName, usecols=[0, 1, 2, 4, 5])
    # print(data)
    keta = int(math.log10(data["n (cm-3)"].max()))
    data["n (cm-3)"] = data["n (cm-3)"] / 10**keta
    print(keta)
    with pd.ExcelWriter(xlFileName + ".xlsx", engine="openpyxl", mode='a') as writer:
        data.to_excel(writer, sheet_name=tsvFileName, index=False)

    # エクセル上でグラフを作成する場合、openpyxlが楽そうなので、
    # ここからopenpyxlを使います。
    wb = load_workbook(xlFileName + ".xlsx")
    sheetNames = wb.sheetnames
    ws = wb[sheetNames[i]]
    maxRow = ws.max_row


    # グラフを２種類用意
    # 後で合体させる
    # この後、設定の順序によってはエラーが出るので注意
    chart1 = ScatterChart()
    chart1.height = chartHeight
    chart1.width = chartWidth
    chart2 = ScatterChart()
    chart2.height = chartHeight
    chart2.width = chartWidth
    # データを準備
    # 各データについて
    #   Referenceオブジェクト -> Seriesオブジェクト -> chartに追加
    # を行う
    # (col, row) = (1, 1) -> A1
    #            = (1, 2) -> B1
    # データ系列に対する書式設定はここから
    # 単純なデータ配置ならadd_dataの方が楽
    #region
    # series.spPr = GraphicalProperties(ln =LineProperties(solidFill="000000", w=1*12700, prstDash="dot"))
    xValue = Reference(ws, min_col=1, min_row=2, max_row=maxRow)
    EcValue = Reference(ws, min_col=2, min_row=2, max_row=maxRow)
    Ec = Series(EcValue, xValue, title = "Ec")
    Ec.spPr.ln.w = 4 * 12700
    EvValue = Reference(ws, min_col=3, min_row=2, max_row=maxRow)
    Ev = Series(EvValue, xValue, title = "Ev")
    Ev.spPr.ln.w = 4 * 12700
    EfValue = Reference(ws, min_col=4, min_row=2, max_row=maxRow)
    Ef = Series(EfValue, xValue, title = "Ef")
    Ef.spPr.ln.w = 4 * 12700
    nValue = Reference(ws, min_col=5, min_row=2, max_row=maxRow)
    n = Series(nValue, xValue, title  ="n")
    n.spPr.ln.w = 4 * 12700
    #endregion
    chart1.series.append(n)
    chart2.series.append(Ef)
    chart2.series.append(Ec)
    chart2.series.append(Ev)

    # グラフ全体に対する書式設定はここから
    #region
    # タイトルの設定
    # chart1.title = "バンドギャップ線図"
    # xchart1.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=2800)
    chart1.x_axis.title = "表面からの深さ [Å]"
    # ↓openpyxlでは、複雑なフォントの設定は無理っぽい。RichTextProperties使えば多少はいける。
    chart1.y_axis.title = "電子濃度 [✕10^"+ str(keta) +" cm-3]" 
    chart2.y_axis.axId = 200
    chart2.y_axis.title = "エネルギー [eV]"
    #グリッド線削除
    chart1.x_axis.majorGridlines = None
    chart1.y_axis.majorGridlines = None
    chart2.y_axis.majorGridlines = None


    font = Font(typeface='Verdana')
    size = 2800 # 28 pt
    char_prop = CharacterProperties(latin=font, sz=size, b=False)
    para_prop = ParagraphProperties(defRPr=char_prop)
    rich_text = RichText(p=[Paragraph(pPr=para_prop, endParaRPr=char_prop)])
    chart1.x_axis.txPr = rich_text
    chart1.y_axis.txPr = rich_text
    chart2.y_axis.txPr = rich_text
    char_prop = CharacterProperties(latin=font, sz=size, b=True, solidFill="000000")
    para_prop = ParagraphProperties(defRPr=char_prop)
    chart1.x_axis.title.tx.rich.p[0].pPr = para_prop
    chart1.y_axis.title.tx.rich.p[0].pPr = para_prop
    chart2.y_axis.title.tx.rich.p[0].pPr = para_prop

    # 軸
    chart1.x_axis.spPr = GraphicalProperties(ln =LineProperties(solidFill="000000", w=4*12700)) # w : 軸の幅
    chart1.y_axis.spPr = GraphicalProperties(ln =LineProperties(solidFill="000000", w=4*12700))
    chart2.y_axis.spPr = GraphicalProperties(ln =LineProperties(solidFill="000000", w=4*12700))
    chart1.x_axis.scaling.min = 0
    chart1.x_axis.scaling.max = 200
    chart1.x_axis.majorTickMark = "in"
    chart1.y_axis.majorTickMark = "in"
    chart2.y_axis.majorTickMark = "in"
    chart1.y_axis.numFmt = "0.0"

    chart1.plot_area.spPr = GraphicalProperties(ln = LineProperties(solidFill = "000000", w=4*12700))

    # 凡例の設定
    # chart1.legend = None
    chart1.legend.legendPos = "t" # l, r, t, b, trから選択
    rich_text = RichText(p=[Paragraph(pPr=para_prop, endParaRPr=char_prop)])
    chart1.legend.txPr = rich_text

    chart1.y_axis.crosses  = "max"
    # chart1.layout = Layout(ManualLayout(x = 0, y = 0, h = 0.8, w = 0.8))
    chart1 += chart2
    ws.add_chart(chart1, "D4")
    #endregion
    i += 1
    wb.save(xlFileName + ".xlsx")
